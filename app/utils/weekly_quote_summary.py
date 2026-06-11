from __future__ import annotations

import os
import shutil
from datetime import date, timedelta
from collections import OrderedDict
from decimal import Decimal, InvalidOperation, ROUND_FLOOR
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

SUPPLIERS = ("勾庄", "理想", "刘慧", "酱菜", "豆制品")
BUILTIN_SUPPLIERS = SUPPLIERS
HIGHEST_PRICE_SUPPLIERS = {"勾庄", "刘慧", "酱菜", "豆制品"}
SUPPLIER_BATCH_LIMITS = {
    "勾庄": 7,
    "理想": 1,
    "刘慧": 1,
    "酱菜": 7,
    "豆制品": 7,
}
DEFAULT_SUPPLIER_CONFIGS = tuple(
    {
        "name": supplier,
        "weekly_batch_limit": SUPPLIER_BATCH_LIMITS[supplier],
        "summary_rule": "highest" if supplier in HIGHEST_PRICE_SUPPLIERS else "average",
        "is_builtin": True,
        "sort_order": index * 10,
    }
    for index, supplier in enumerate(SUPPLIERS, start=1)
)
IMPORT_REQUIRED_COLUMNS = ("菜名", "单位", "单价")
OUTPUT_HEADERS = ("菜名", "单位", "汇总价")
READABLE_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}
WRITABLE_EXTENSIONS = {".xlsx", ".xlsm"}


def import_weekly_quote_batch(
    source_path: str,
    supplier: str,
    quote_date: str,
    supplier_configs: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    configs = _normalize_supplier_configs(supplier_configs)
    normalized_supplier = _normalize_supplier(supplier, configs)
    normalized_date = _normalize_text(quote_date)
    if not normalized_date:
        raise ValueError("报价日期不能为空")
    if not source_path or not os.path.exists(source_path):
        raise FileNotFoundError(f"导入文件不存在: {source_path}")

    data_frame = _read_import_excel(source_path, normalized_supplier)
    column_map = _resolve_import_columns(data_frame.columns)

    entries: list[dict[str, Any]] = []
    for row_offset, (_, row) in enumerate(data_frame.iterrows(), start=2):
        raw_name = row.iloc[column_map["菜名"]]
        raw_unit = row.iloc[column_map["单位"]]
        raw_price = row.iloc[column_map["单价"]]

        name = _normalize_text(raw_name)
        unit = _normalize_text(raw_unit)
        price_text = _normalize_text(raw_price)
        if not name and not unit and not price_text:
            continue
        if not name:
            raise ValueError(f"第 {row_offset} 行缺少菜名")
        if not unit:
            raise ValueError(f"第 {row_offset} 行缺少单位")
        price = _normalize_price(raw_price, f"第 {row_offset} 行")
        entries.append(
            {
                "name": name,
                "unit": unit,
                "price": float(price),
            }
        )

    if not entries:
        raise ValueError("导入文件没有可用的报价记录")

    return {
        "supplier": normalized_supplier,
        "quote_date": normalized_date,
        "entries": entries,
    }


def preview_weekly_quote_summary(
    batches: list[dict[str, Any]],
    supplier_configs: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    configs = _normalize_supplier_configs(supplier_configs)
    normalized_batches = _normalize_batches(batches, configs)
    return _build_preview_from_batches(normalized_batches, configs)


def _build_preview_from_batches(
    normalized_batches: list[dict[str, Any]],
    supplier_configs: list[dict[str, Any]],
) -> dict[str, Any]:
    unit_summaries = _build_unit_summaries(normalized_batches, supplier_configs)
    total_entries = sum(summary["entry_count"] for summary in unit_summaries)
    total_summary_items = sum(len(summary["summary_items"]) for summary in unit_summaries)

    return {
        "unit_summaries": unit_summaries,
        "total_batches": len(normalized_batches),
        "total_entries": total_entries,
        "total_summary_items": total_summary_items,
        "issue_messages": [],
    }


def export_weekly_quote_summary(
    workbook_path: str,
    batches: list[dict[str, Any]],
    supplier_configs: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not workbook_path or not os.path.exists(workbook_path):
        raise FileNotFoundError(f"输出工作簿不存在: {workbook_path}")

    workbook_ext = os.path.splitext(workbook_path)[1].lower()
    if workbook_ext not in WRITABLE_EXTENSIONS:
        supported = ", ".join(sorted(WRITABLE_EXTENSIONS))
        raise ValueError(f"总结表只支持这些格式: {supported}")

    configs = _normalize_supplier_configs(supplier_configs)
    supplier_names = [config["name"] for config in configs]
    normalized_batches = _normalize_batches(batches, configs)
    preview = _build_preview_from_batches(normalized_batches, configs)
    unit_summaries = preview["unit_summaries"]
    summary_map = {
        summary["supplier"]: summary
        for summary in unit_summaries
    }
    output_path = _build_summary_output_path(workbook_path, normalized_batches)
    _copy_template_workbook(workbook_path, output_path)

    workbook = load_workbook(output_path, keep_vba=workbook_ext == ".xlsm")
    resolved_sheet_names = _resolve_supplier_sheet_names(workbook.sheetnames, supplier_names)

    for supplier in supplier_names:
        summary = summary_map[supplier]
        worksheet = _prepare_summary_sheet(
            workbook=workbook,
            sheet_name=resolved_sheet_names.get(supplier, supplier),
        )
        _write_summary_sheet(worksheet, summary["summary_items"])

    try:
        workbook.save(output_path)
    except PermissionError as exc:
        raise PermissionError(f"无法写入总结表，请关闭已打开的文件: {output_path}") from exc

    return {
        "workbook_path": output_path,
        "sheet_names": [resolved_sheet_names.get(supplier, supplier) for supplier in supplier_names],
        "unit_summaries": unit_summaries,
        "total_batches": preview["total_batches"],
        "total_entries": preview["total_entries"],
        "total_summary_items": preview["total_summary_items"],
    }


def _read_import_excel(source_path: str, supplier: str) -> pd.DataFrame:
    ext = os.path.splitext(source_path)[1].lower()
    if ext not in READABLE_EXTENSIONS:
        readable = ", ".join(sorted(READABLE_EXTENSIONS))
        raise ValueError(f"导入文件只支持这些 Excel 格式: {readable}")

    if ext == ".xls":
        try:
            workbook = pd.ExcelFile(source_path, engine="xlrd")
        except ImportError as exc:
            raise ValueError("当前环境缺少 .xls 读取支持，请改用 .xlsx 文件") from exc
    else:
        workbook = pd.ExcelFile(source_path, engine="openpyxl")

    try:
        sheet_name = _match_supplier_sheet_name(workbook.sheet_names, supplier)
        return workbook.parse(sheet_name=sheet_name)
    finally:
        workbook.close()


def _resolve_import_columns(columns: Any) -> dict[str, int]:
    normalized_headers = {_normalize_text(name): index for index, name in enumerate(columns)}
    missing_headers = [name for name in IMPORT_REQUIRED_COLUMNS if name not in normalized_headers]
    if missing_headers:
        raise ValueError(f"导入模板缺少必要列: {'、'.join(missing_headers)}")
    return {name: normalized_headers[name] for name in IMPORT_REQUIRED_COLUMNS}


def _normalize_batches(
    batches: list[dict[str, Any]],
    supplier_configs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not batches:
        raise ValueError("当前没有可汇总的报价批次")

    normalized_batches: list[dict[str, Any]] = []
    supplier_names = [config["name"] for config in supplier_configs]
    supplier_counts = {supplier: 0 for supplier in supplier_names}
    supplier_limits = {config["name"]: int(config["weekly_batch_limit"]) for config in supplier_configs}

    for batch_index, batch in enumerate(batches, start=1):
        supplier = _normalize_supplier(batch.get("supplier"), supplier_configs)
        supplier_counts[supplier] += 1
        limit = supplier_limits[supplier]
        if supplier_counts[supplier] > limit:
            raise ValueError(f"{supplier} 本周最多只允许 {limit} 个批次")

        quote_date = _normalize_text(batch.get("quote_date"))
        if not quote_date:
            raise ValueError(f"第 {batch_index} 个批次缺少报价日期")

        raw_entries = batch.get("entries") or []
        normalized_entries: list[dict[str, Any]] = []
        for entry_index, entry in enumerate(raw_entries, start=1):
            name = _normalize_text(entry.get("name"))
            unit = _normalize_text(entry.get("unit"))
            raw_price = entry.get("price")
            if not name and not unit and _is_blank(raw_price):
                continue
            if not name:
                raise ValueError(f"{supplier} {quote_date} 第 {entry_index} 条记录缺少菜名")
            if not unit:
                raise ValueError(f"{supplier} {quote_date} 第 {entry_index} 条记录缺少单位")
            price = _normalize_price(raw_price, f"{supplier} {quote_date} 第 {entry_index} 条记录")
            normalized_entries.append(
                {
                    "name": name,
                    "unit": unit,
                    "price": price,
                }
            )

        if not normalized_entries:
            # 空批次静默跳过（可能是用户清空了所有记录后保存导致）
            continue

        normalized_batches.append(
            {
                "supplier": supplier,
                "quote_date": quote_date,
                "entries": normalized_entries,
            }
        )

    return normalized_batches


def _build_unit_summaries(
    batches: list[dict[str, Any]],
    supplier_configs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    supplier_names = [config["name"] for config in supplier_configs]
    supplier_rules = {config["name"]: config["summary_rule"] for config in supplier_configs}
    supplier_batch_counts = {supplier: 0 for supplier in supplier_names}
    supplier_entry_counts = {supplier: 0 for supplier in supplier_names}
    supplier_buckets: dict[str, OrderedDict[tuple[str, str], dict[str, Any]]] = {
        supplier: OrderedDict()
        for supplier in supplier_names
    }

    for batch in batches:
        supplier = batch["supplier"]
        supplier_batch_counts[supplier] += 1
        supplier_entry_counts[supplier] += len(batch["entries"])

        for entry in batch["entries"]:
            bucket_key = (entry["name"], entry["unit"])
            bucket = supplier_buckets[supplier].setdefault(
                bucket_key,
                {
                    "name": entry["name"],
                    "unit": entry["unit"],
                    "prices": [],
                },
            )
            bucket["prices"].append(entry["price"])

    unit_summaries: list[dict[str, Any]] = []
    for supplier in supplier_names:
        summary_items = []
        for bucket in supplier_buckets[supplier].values():
            summary_price = _summarize_prices_for_supplier(supplier, bucket["prices"], supplier_rules)
            summary_items.append(
                {
                    "name": bucket["name"],
                    "unit": bucket["unit"],
                    "summary_price": float(summary_price),
                    # Backward-compatible legacy field for older API consumers.
                    "average_price": float(summary_price),
                }
            )

        unit_summaries.append(
            {
                "supplier": supplier,
                "batch_count": supplier_batch_counts[supplier],
                "entry_count": supplier_entry_counts[supplier],
                "summary_items": summary_items,
            }
        )

    return unit_summaries


def _summarize_prices_for_supplier(
    supplier: str,
    prices: list[Decimal],
    supplier_rules: dict[str, str],
) -> Decimal:
    if supplier_rules.get(supplier) == "highest":
        return max(prices)
    return _average_prices(prices)


def _average_prices(prices: list[Decimal]) -> Decimal:
    total = sum(prices, Decimal("0"))
    average = total / Decimal(len(prices))
    return _round_two_drop_three_up(average)


def _round_two_drop_three_up(value: Decimal) -> Decimal:
    base = (value * Decimal("10")).to_integral_value(rounding=ROUND_FLOOR) / Decimal("10")
    remainder = value - base
    if remainder >= Decimal("0.03"):
        return base + Decimal("0.1")
    return base


def _normalize_supplier_configs(
    supplier_configs: list[dict[str, Any]] | None,
) -> list[dict[str, Any]]:
    raw_configs = supplier_configs or list(DEFAULT_SUPPLIER_CONFIGS)
    configs: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, config in enumerate(raw_configs, start=1):
        name = _normalize_text(config.get("name") or config.get("supplier"))
        if not name or name in seen:
            continue
        rule = _normalize_text(config.get("summary_rule") or _default_summary_rule_for_supplier(name))
        if rule not in {"highest", "average"}:
            rule = _default_summary_rule_for_supplier(name)
        try:
            limit = int(config.get("weekly_batch_limit") or config.get("limit") or SUPPLIER_BATCH_LIMITS.get(name, 7))
        except (TypeError, ValueError):
            limit = SUPPLIER_BATCH_LIMITS.get(name, 7)
        limit = max(1, min(7, limit))
        configs.append(
            {
                "name": name,
                "weekly_batch_limit": limit,
                "summary_rule": rule,
                "is_builtin": bool(config.get("is_builtin", name in SUPPLIERS)),
                "sort_order": int(config.get("sort_order") or index * 10),
            }
        )
        seen.add(name)
    return configs


def _default_summary_rule_for_supplier(supplier: str) -> str:
    return "highest" if supplier in HIGHEST_PRICE_SUPPLIERS else "average"


def _normalize_supplier(supplier: Any, supplier_configs: list[dict[str, Any]]) -> str:
    value = _normalize_text(supplier)
    allowed = {config["name"] for config in supplier_configs}
    if value not in allowed:
        raise ValueError(f"不支持的单位: {supplier}")
    return value


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    return text


def _normalize_price(raw_value: Any, label: Any) -> Decimal:
    if _is_blank(raw_value):
        raise ValueError(f"{label} 缺少单价")
    try:
        price = Decimal(str(raw_value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{label} 的单价不是有效数字") from exc
    if price <= 0:
        raise ValueError(f"{label} 的单价必须大于 0")
    return price


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    return str(value).strip() == ""


def _normalize_sheet_key(value: Any) -> str:
    text = _normalize_text(value)
    return "".join(ch for ch in text.lower() if ch.isalnum())


def _sheet_match_score(sheet_name: str, supplier: str) -> tuple[int, int]:
    supplier_key = _normalize_sheet_key(supplier)
    sheet_key = _normalize_sheet_key(sheet_name)
    if not supplier_key or not sheet_key:
        return (0, 0)
    if sheet_key == supplier_key:
        return (300, -len(sheet_key))
    if supplier_key in sheet_key:
        starts_or_ends = 20 if (sheet_key.startswith(supplier_key) or sheet_key.endswith(supplier_key)) else 0
        return (200 + starts_or_ends, -len(sheet_key))
    return (0, 0)


def _match_supplier_sheet_name(sheet_names: list[str], supplier: str) -> str:
    scored_matches = []
    for sheet_name in sheet_names:
        score = _sheet_match_score(sheet_name, supplier)
        if score[0] > 0:
            scored_matches.append((score, sheet_name))

    if not scored_matches:
        raise ValueError(
            f"导入模板中找不到“{supplier}”对应的 sheet。当前可用 sheet: {'、'.join(sheet_names)}"
        )

    scored_matches.sort(key=lambda item: item[0], reverse=True)
    best_score = scored_matches[0][0]
    best_matches = [sheet_name for score, sheet_name in scored_matches if score == best_score]
    if len(best_matches) > 1:
        raise ValueError(
            f"导入模板中有多个 sheet 都像“{supplier}”：{'、'.join(best_matches)}，请调整 sheet 名称后再导入"
        )
    return best_matches[0]


def _resolve_supplier_sheet_names(sheet_names: list[str], suppliers: list[str]) -> dict[str, str]:
    resolved: dict[str, str] = {}
    remaining_sheet_names = list(sheet_names)

    for supplier in suppliers:
        try:
            matched_sheet = _match_supplier_sheet_name(remaining_sheet_names, supplier)
        except ValueError:
            continue
        resolved[supplier] = matched_sheet
        remaining_sheet_names.remove(matched_sheet)

    return resolved


def _prepare_summary_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        max_row = max(worksheet.max_row, 1)
        for row_index in range(1, max_row + 1):
            for column_index in range(1, 4):
                worksheet.cell(row=row_index, column=column_index, value=None)
        return worksheet
    return workbook.create_sheet(title=sheet_name)


def _format_week_label(value: date) -> str:
    return f"{value.year % 100}.{value.month}.{value.day}"


def _resolve_output_base_date(_normalized_batches: list[dict[str, Any]]) -> date:
    return date.today()


def _build_summary_output_path(
    template_path: str,
    normalized_batches: list[dict[str, Any]],
) -> str:
    base_date = _resolve_output_base_date(normalized_batches)
    days_until_next_monday = 7 - base_date.weekday()
    next_week_start = base_date + timedelta(days=days_until_next_monday)
    next_week_end = next_week_start + timedelta(days=6)
    template_ext = os.path.splitext(template_path)[1].lower()
    filename = (
        f"{_format_week_label(next_week_start)}-"
        f"{_format_week_label(next_week_end)}-每周报价总结{template_ext}"
    )
    return os.path.join(os.path.dirname(template_path), filename)


def _copy_template_workbook(template_path: str, output_path: str) -> None:
    try:
        shutil.copy2(template_path, output_path)
    except PermissionError as exc:
        raise PermissionError(f"无法生成新的总结表，请关闭已打开的文件: {output_path}") from exc


def _write_summary_sheet(worksheet, summary_items: list[dict[str, Any]]) -> None:
    worksheet.cell(row=1, column=1, value=OUTPUT_HEADERS[0])
    worksheet.cell(row=1, column=2, value=OUTPUT_HEADERS[1])
    worksheet.cell(row=1, column=3, value=OUTPUT_HEADERS[2])

    for column_index in range(1, 4):
        worksheet.cell(row=1, column=column_index).font = Font(bold=True)

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 12

    for row_index, item in enumerate(summary_items, start=2):
        worksheet.cell(row=row_index, column=1, value=item["name"])
        worksheet.cell(row=row_index, column=2, value=item["unit"])
        worksheet.cell(row=row_index, column=3, value=float(item["summary_price"]))
