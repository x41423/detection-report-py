import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.utils.weekly_quote_summary import (
    export_weekly_quote_summary,
    import_weekly_quote_batch,
    preview_weekly_quote_summary,
)


class FakeSystemDate(__import__("datetime").date):
    @classmethod
    def today(cls):
        return cls(2026, 4, 18)


class WeeklyQuoteSummaryTests(unittest.TestCase):
    def test_import_batch_uses_supplier_sheet_instead_of_first_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "quotes.xlsx"

            with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                pd.DataFrame(
                    {
                        "菜名": ["错误菜"],
                        "单位": ["斤"],
                        "单价": [99],
                    }
                ).to_excel(writer, sheet_name="理想报价", index=False)
                pd.DataFrame(
                    {
                        "菜名": ["白菜", "萝卜"],
                        "单位": ["斤", "斤"],
                        "单价": [2.5, 1.8],
                    }
                ).to_excel(writer, sheet_name="02-勾庄", index=False)

            batch = import_weekly_quote_batch(
                source_path=str(workbook_path),
                supplier="勾庄",
                quote_date="2026-04-18",
            )

            self.assertEqual(batch["supplier"], "勾庄")
            self.assertEqual(len(batch["entries"]), 2)
            self.assertEqual(batch["entries"][0]["name"], "白菜")
            self.assertEqual(batch["entries"][1]["name"], "萝卜")

    def test_export_summary_copies_template_with_next_week_filename(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "summary-template.xlsx"

            workbook = Workbook()
            workbook.active.title = "豆制品模板"
            for sheet_name in ["说明", "刘慧报价", "勾庄模板", "酱菜汇总", "理想-模板"]:
                workbook.create_sheet(title=sheet_name)

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                worksheet["A1"] = "模板表头"
                worksheet["A2"] = "模板原值"
                worksheet["D1"] = "保留列"

            workbook.save(template_path)

            with patch("app.utils.weekly_quote_summary.date", FakeSystemDate):
                result = export_weekly_quote_summary(
                    workbook_path=str(template_path),
                    batches=[
                        {
                            "supplier": "勾庄",
                            "quote_date": "2026-04-18",
                            "entries": [
                                {"name": "白菜", "unit": "斤", "price": 2.5},
                                {"name": "白菜", "unit": "斤", "price": 1.5},
                            ],
                        }
                    ],
                )

            output_path = Path(result["workbook_path"])

            self.assertTrue(output_path.exists())
            self.assertNotEqual(output_path, template_path)
            self.assertEqual(output_path.name, "26.4.20-26.4.26-每周报价总结.xlsx")
            self.assertIn("勾庄模板", result["sheet_names"])

            template_workbook = load_workbook(template_path)
            output_workbook = load_workbook(output_path)

            self.assertEqual(template_workbook["勾庄模板"]["A2"].value, "模板原值")
            self.assertEqual(output_workbook["勾庄模板"]["A1"].value, "菜名")
            self.assertEqual(output_workbook["勾庄模板"]["A2"].value, "白菜")
            self.assertEqual(output_workbook["勾庄模板"]["B2"].value, "斤")
            self.assertEqual(output_workbook["勾庄模板"]["C2"].value, 2.5)
            self.assertEqual(output_workbook["勾庄模板"]["D1"].value, "保留列")

    def test_preview_uses_highest_price_for_specific_suppliers_only(self):
        summary = preview_weekly_quote_summary(
            [
                {
                    "supplier": "勾庄",
                    "quote_date": "2026-04-18",
                    "entries": [
                        {"name": "白菜", "unit": "斤", "price": 2.5},
                        {"name": "白菜", "unit": "斤", "price": 1.5},
                    ],
                },
                {
                    "supplier": "酱菜",
                    "quote_date": "2026-04-18",
                    "entries": [
                        {"name": "萝卜干", "unit": "包", "price": 5.2},
                        {"name": "萝卜干", "unit": "包", "price": 4.8},
                    ],
                },
                {
                    "supplier": "豆制品",
                    "quote_date": "2026-04-18",
                    "entries": [
                        {"name": "千张", "unit": "斤", "price": 6.0},
                        {"name": "千张", "unit": "斤", "price": 6.3},
                    ],
                },
                {
                    "supplier": "理想",
                    "quote_date": "2026-04-18",
                    "entries": [
                        {"name": "青菜", "unit": "斤", "price": 2.5},
                        {"name": "青菜", "unit": "斤", "price": 1.5},
                    ],
                },
            ]
        )

        unit_summaries = {item["supplier"]: item for item in summary["unit_summaries"]}

        self.assertEqual(unit_summaries["勾庄"]["summary_items"][0]["summary_price"], 2.5)
        self.assertEqual(unit_summaries["酱菜"]["summary_items"][0]["summary_price"], 5.2)
        self.assertEqual(unit_summaries["豆制品"]["summary_items"][0]["summary_price"], 6.3)
        self.assertEqual(unit_summaries["理想"]["summary_items"][0]["summary_price"], 2.0)
        self.assertEqual(unit_summaries["勾庄"]["summary_items"][0]["average_price"], 2.5)
        self.assertEqual(unit_summaries["理想"]["summary_items"][0]["average_price"], 2.0)


if __name__ == "__main__":
    unittest.main()
