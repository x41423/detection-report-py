import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.services.doc_service import DocService
from backend.services.monthly_list_parser import MonthlyListParser
from backend.services.pesticide_service import PesticideService
from backend.services.template_library_service import (
    get_pesticide_templates,
    save_pesticide_template,
    save_transfer_template,
)


class MonthlyBatchTests(unittest.TestCase):
    def _save_formula_monthly_list(self, path: Path, columns: int, filled_columns: int) -> None:
        workbook = Workbook()
        sheet = workbook.active
        for column in range(1, columns + 1):
            letter = get_column_letter(column)
            sheet.cell(row=1, column=column).value = (
                f'=IF(COLUMN({letter}1)<=DAY(EOMONTH(TODAY(),0)),'
                f'TEXT(DATE(YEAR(TODAY()),MONTH(TODAY()),COLUMN({letter}1)),"yyyy:mm:dd"),"")'
            )
            if column <= filled_columns:
                sheet.cell(row=2, column=column).value = f"菜{column}"
        workbook.save(path)

    def test_monthly_list_parser_parses_column_format_and_splits_cell_names(self):
        parser = MonthlyListParser()

        result = parser.parse_text(
            "2026.4.1\t2026.4.2\n白菜、黄瓜\t菠菜\n萝卜\t芹菜\n白菜\t",
            "2026-04",
        )

        self.assertEqual(result["total_dates"], 2)
        self.assertEqual(result["total_names"], 5)
        self.assertEqual(result["entries"][0]["date"], "2026-04-01")
        self.assertEqual(result["entries"][0]["names"], ["白菜", "黄瓜", "萝卜"])
        self.assertEqual(result["entries"][1]["date"], "2026-04-02")
        self.assertEqual(result["entries"][1]["names"], ["菠菜", "芹菜"])
        self.assertEqual(result["errors"], [])

    def test_monthly_list_parser_accepts_colon_date_headers(self):
        parser = MonthlyListParser()

        result = parser.parse_text("2026:04:01\t2026:04:02\n白菜\t黄瓜", "2026-04")

        self.assertEqual(result["total_dates"], 2)
        self.assertEqual([entry["date"] for entry in result["entries"]], ["2026-04-01", "2026-04-02"])
        self.assertEqual(result["errors"], [])

    def test_monthly_list_parser_parses_xlsx_columns(self):
        parser = MonthlyListParser()

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "月度清单.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "2026.4.1"
            sheet["B1"] = "2026.4.2"
            sheet["A2"] = "白菜\n黄瓜"
            sheet["B2"] = "菠菜、芹菜"
            workbook.save(path)

            result = parser.parse_file(path, "2026-04")

        self.assertEqual(result["total_dates"], 2)
        self.assertEqual(result["entries"][0]["names"], ["白菜", "黄瓜"])
        self.assertEqual(result["entries"][1]["names"], ["菠菜", "芹菜"])

    def test_monthly_list_parser_derives_xlsx_formula_dates_without_cached_values(self):
        parser = MonthlyListParser()

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "公式月度清单.xlsx"
            self._save_formula_monthly_list(path, columns=3, filled_columns=3)

            result = parser.parse_file(path, "2026-04")

        self.assertEqual([entry["date"] for entry in result["entries"]], [
            "2026-04-01",
            "2026-04-02",
            "2026-04-03",
        ])
        self.assertEqual(result["entries"][0]["names"], ["菜1"])
        self.assertEqual(result["errors"], [])

    def test_monthly_list_parser_respects_month_lengths_for_formula_columns(self):
        parser = MonthlyListParser()

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "公式月度清单.xlsx"

            self._save_formula_monthly_list(path, columns=31, filled_columns=31)
            january = parser.parse_file(path, "2026-01")
            self.assertEqual(january["total_dates"], 31)
            self.assertEqual(january["entries"][-1]["date"], "2026-01-31")
            self.assertEqual(january["errors"], [])

            self._save_formula_monthly_list(path, columns=31, filled_columns=30)
            april_without_31 = parser.parse_file(path, "2026-04")
            self.assertEqual(april_without_31["total_dates"], 30)
            self.assertEqual(april_without_31["errors"], [])

            self._save_formula_monthly_list(path, columns=31, filled_columns=31)
            april_with_31 = parser.parse_file(path, "2026-04")
            self.assertEqual(april_with_31["total_dates"], 30)
            self.assertEqual(len(april_with_31["errors"]), 1)
            self.assertIn("没有第 31 天", april_with_31["errors"][0]["message"])

            self._save_formula_monthly_list(path, columns=29, filled_columns=29)
            february_non_leap = parser.parse_file(path, "2025-02")
            self.assertEqual(february_non_leap["total_dates"], 28)
            self.assertEqual(len(february_non_leap["errors"]), 1)
            self.assertIn("没有第 29 天", february_non_leap["errors"][0]["message"])

            february_leap = parser.parse_file(path, "2024-02")
            self.assertEqual(february_leap["total_dates"], 29)
            self.assertEqual(february_leap["entries"][-1]["date"], "2024-02-29")
            self.assertEqual(february_leap["errors"], [])

    def test_monthly_list_parser_reports_column_errors_and_legacy_hint(self):
        parser = MonthlyListParser()

        result = parser.parse_text("2026.4.1\t2026.5.1\t\n白菜\t萝卜\t黄瓜", "2026-04")

        self.assertEqual(result["total_dates"], 2)
        self.assertEqual(result["detected_month"], "2026-04")
        self.assertEqual(len(result["errors"]), 1)
        self.assertIn("首行未识别到日期", result["errors"][0]["message"])

        legacy_result = parser.parse_text("日期\t菜名\n2026-04-01\t白菜、黄瓜", "2026-04")
        self.assertEqual(legacy_result["total_dates"], 0)
        self.assertIn("每列一天", legacy_result["errors"][0]["message"])

    def test_template_library_saves_pesticide_and_transfer_templates(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source = root / "模板.docx"
            source.write_bytes(b"template")
            config: dict = {
                "pesticide_templates": {},
                "transfer_templates": {},
                "small_templates": {},
            }

            def fake_get_config():
                return dict(config)

            def fake_update_config(updates):
                config.update(updates)
                return dict(config)

            with patch("backend.services.template_library_service._template_root", return_value=root / "templates"), patch(
                "backend.services.template_library_service.get_config", side_effect=fake_get_config
            ), patch("backend.services.template_library_service.update_config", side_effect=fake_update_config):
                pesticide_status = save_pesticide_template("big", source, "农残大表模板.docx")
                transfer_status = save_transfer_template("滨鲜", source, "滨鲜模板.docx")
                queried = get_pesticide_templates()

            self.assertTrue(pesticide_status["big_template"]["configured"])
            self.assertTrue(queried["big_template"]["configured"])
            self.assertTrue(transfer_status["templates"]["滨鲜"]["configured"])
            self.assertTrue(Path(config["small_templates"]["滨鲜"]).exists())

    def test_doc_service_groups_monthly_tables_by_filename_date(self):
        service = DocService()
        paths = [
            r"C:\tmp\农残检测记录表2026.04.14.docx",
            r"C:\tmp\农残检测记录表2026.04.14-1.docx",
            r"C:\tmp\农残检测记录表2026.04.15.docx",
            r"C:\tmp\其他文件.docx",
        ]

        result = service.preview_monthly_groups(paths, "2026-04")

        self.assertEqual([group["date"] for group in result["groups"]], ["2026-04-14", "2026-04-15"])
        self.assertEqual(result["groups"][0]["count"], 2)
        self.assertEqual(len(result["unrecognized_files"]), 1)

    def test_pesticide_monthly_execute_generates_manifest_and_flat_outputs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            big_template = root / "big.docx"
            small_template = root / "small.docx"
            output_dir = root / "output"
            big_template.write_bytes(b"big")
            small_template.write_bytes(b"small")
            output_dir.mkdir()

            def fake_process_documents(big_path, small_path, data, date_label, output_dir_text, inspector_name):
                output_root = Path(output_dir_text)
                Path(big_path).with_suffix(".docx")
                output_root.joinpath(Path(big_path).name).write_bytes(b"big-out")
                output_root.joinpath(Path(small_path).name).write_bytes(b"small-out")

            with patch("backend.services.pesticide_service.get_config", return_value={}), patch(
                "backend.services.pesticide_service.process_documents", side_effect=fake_process_documents
            ), patch.object(
                PesticideService,
                "generate_rates",
                return_value=[{"variety": "白菜", "rate": "6.150%"}],
            ):
                service = PesticideService()
                result = service.execute_monthly_task(
                    [{"date": "2026-04-14", "names": ["白菜"]}],
                    str(big_template),
                    str(small_template),
                    "2026-04",
                    str(output_dir),
                    "检查员",
                )

            self.assertEqual(result["success_count"], 1)
            self.assertTrue((output_dir / "农残检测记录表2026.04.14.docx").exists())
            self.assertTrue((output_dir / "单位农残记录表04.14.docx").exists())
            manifest = json.loads((output_dir / "处理结果清单.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["success_count"], 1)

    def test_pesticide_monthly_execute_deduplicates_names(self):
        """执行月度任务时同名蔬菜应去重，generate_rates 只收到唯一菜名。"""
        captured_names: list[list[str]] = []

        def fake_generate_rates(veg_text: str):
            captured_names.append(veg_text.split("\n"))
            return [{"variety": n, "rate": "6.150%"} for n in veg_text.split("\n") if n]

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            big_template = root / "big.docx"
            small_template = root / "small.docx"
            output_dir = root / "output"
            big_template.write_bytes(b"big")
            small_template.write_bytes(b"small")
            output_dir.mkdir()

            def fake_process_documents(*_args, **_kwargs):
                pass

            with patch("backend.services.pesticide_service.get_config", return_value={}), \
                 patch("backend.services.pesticide_service.process_documents", side_effect=fake_process_documents), \
                 patch.object(PesticideService, "generate_rates", side_effect=fake_generate_rates):
                service = PesticideService()
                service.execute_monthly_task(
                    [{"date": "2026-04-14", "names": ["白菜", "白菜", "菠菜", "白菜"]}],
                    str(big_template),
                    str(small_template),
                    "2026-04",
                    str(output_dir),
                    "检查员",
                )

            self.assertEqual(len(captured_names), 1)
            self.assertEqual(captured_names[0], ["白菜", "菠菜"])

    def test_monthly_list_parser_handles_array_formula_first_column(self):
        """A1 为 ArrayFormula（真实 Excel 中常见），解析后 A 列日期不应被跳过。"""
        from openpyxl import Workbook
        from openpyxl.worksheet.formula import ArrayFormula

        parser = MonthlyListParser()

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "array_formula_list.xlsx"
            wb = Workbook()
            ws = wb.active
            # 设置 A1 为数组公式，覆盖 A1:B1
            ws["A1"] = ArrayFormula(
                "A1:B1",
                '=TEXT(DATE(YEAR(TODAY()),4,COLUMN(A1:B1)),"yyyy:mm:dd")',
            )
            # data_only 模式下 A1 会显示计算结果
            ws["A1"].value = "2026:04:01"
            ws["B1"].value = "2026:04:02"
            ws["A2"] = "白菜"
            ws["B2"] = "菠菜"
            wb.save(path)

            result = parser.parse_file(path, "2026-04")

        self.assertEqual(result["total_dates"], 2)
        self.assertEqual(result["entries"][0]["date"], "2026-04-01")
        self.assertEqual(result["entries"][1]["date"], "2026-04-02")
        self.assertEqual(result["errors"], [])

    def test_pesticide_monthly_execute_uses_confirmed_file_name_format(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            big_template = root / "big.docx"
            small_template = root / "small.docx"
            output_dir = root / "output"
            big_template.write_bytes(b"big")
            small_template.write_bytes(b"small")
            output_dir.mkdir()

            def fake_process_documents(big_path, small_path, data, date_label, output_dir_text, inspector_name):
                output_root = Path(output_dir_text)
                output_root.joinpath(Path(big_path).name).write_bytes(b"big-out")
                output_root.joinpath(Path(small_path).name).write_bytes(b"small-out")

            with patch("backend.services.pesticide_service.get_config", return_value={}), patch(
                "backend.services.pesticide_service.process_documents", side_effect=fake_process_documents
            ):
                service = PesticideService()
                result = service.execute_monthly_task(
                    [{"date": "2026-04-01", "names": ["白菜"]}],
                    str(big_template),
                    str(small_template),
                    "2026-04",
                    str(output_dir),
                    "检查员",
                )

            self.assertEqual(result["success_count"], 1)
            self.assertTrue((output_dir / "农残检测记录表2026.04.01.docx").exists())
            self.assertTrue((output_dir / "单位农残记录表04.1.docx").exists())


if __name__ == "__main__":
    unittest.main()
