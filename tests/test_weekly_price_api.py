import os
import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db.store as store
from backend.main import app
from tests.auth_api_utils import auth_headers_for_permissions


WEEKLY_PRICE_TEST_PERMISSIONS = [
    "weekly_quote:view",
    "weekly_quote:create",
    "weekly_quote:update",
    "weekly_quote:delete",
    "weekly_quote:export",
    "weekly_quote:aliases",
]


class WeeklyPriceApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.original_db_dir = store.DB_DIR
        cls.original_db_path = store.DB_PATH
        store.close_connection()
        store.DB_DIR = cls.temp_dir.name
        store.DB_PATH = os.path.join(cls.temp_dir.name, "weekly-price-api-test.db")
        store._connection = None
        cls.client_ctx = TestClient(app)
        cls.client = cls.client_ctx.__enter__()
        cls.client.headers.update(auth_headers_for_permissions(cls.client, WEEKLY_PRICE_TEST_PERMISSIONS))

    @classmethod
    def tearDownClass(cls):
        cls.client_ctx.__exit__(None, None, None)
        store.close_connection()
        store.DB_DIR = cls.original_db_dir
        store.DB_PATH = cls.original_db_path
        store._connection = None
        cls.temp_dir.cleanup()

    def test_execute_returns_400_when_output_parent_missing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            update_path = root / "update.xlsx"
            reference_path = root / "reference.xlsx"
            update_path.write_text("update", encoding="utf-8")
            reference_path.write_text("reference", encoding="utf-8")

            response = self.client.post(
                "/api/weekly-price/execute",
                json={
                    "update_path": str(update_path),
                    "reference_path": str(reference_path),
                    "output_path": str(root / "missing" / "custom-output.xlsx"),
                },
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("\u8f93\u51fa\u8def\u5f84", response.json()["detail"])

    def test_preview_upload_returns_preview_payload(self):
        with patch(
            "backend.api.routes.weekly_price.service.preview",
            return_value={
                "success": True,
                "message": "preview-ok",
                "matched_count": 3,
                "updated_count": 3,
                "matched_items": [],
                "not_matched": [],
                "not_matched_count": 0,
                "not_matched_unique_count": 0,
                "suggested_matches": [],
                "alias_hit_count": 0,
                "warnings": [],
                "update_start_row": 1,
                "reference_start_row": 1,
            },
        ):
            response = self.client.post(
                "/api/weekly-price/preview/upload",
                files=[
                    (
                        "update_file",
                        (
                            "update.xlsx",
                            b"update",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    ),
                    (
                        "reference_file",
                        (
                            "reference.xlsx",
                            b"reference",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    ),
                ],
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["matched_count"], 3)

    def test_execute_upload_returns_downloadable_file(self):
        def fake_execute(update_path: str, reference_path: str, output_path: str):
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            Path(output_path).write_bytes(b"fake-xlsx")
            return {
                "success": True,
                "message": "execute-ok",
                "matched_count": 4,
                "updated_count": 4,
                "matched_items": [],
                "not_matched": [],
                "not_matched_count": 0,
                "not_matched_unique_count": 0,
                "alias_hit_count": 0,
                "warnings": [],
                "output_path": output_path,
                "backup_path": None,
            }

        with patch("backend.api.routes.weekly_price.service.execute", side_effect=fake_execute):
            response = self.client.post(
                "/api/weekly-price/execute/upload",
                files=[
                    (
                        "update_file",
                        (
                            "update.xlsx",
                            b"update",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    ),
                    (
                        "reference_file",
                        (
                            "reference.xlsx",
                            b"reference",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    ),
                ],
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, b"fake-xlsx")
        self.assertIn("attachment;", response.headers["content-disposition"])
        self.assertEqual(response.headers["x-operation-message"], "execute-ok")
        self.assertEqual(response.headers["x-matched-count"], "4")
        self.assertEqual(response.headers["x-updated-count"], "4")

    def test_preview_endpoint_requires_authentication(self):
        self.client.headers.clear()

        response = self.client.post("/api/weekly-price/preview", json={"update_path": "a", "reference_path": "b"})

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()["detail"]["code"], "AUTH_REQUIRED")
        self.client.headers.update(auth_headers_for_permissions(self.client, WEEKLY_PRICE_TEST_PERMISSIONS))

    def test_execute_endpoint_requires_update_permission(self):
        self.client.headers.clear()
        self.client.headers.update(auth_headers_for_permissions(self.client, ["weekly_quote:view"]))

        response = self.client.post(
            "/api/weekly-price/execute",
            json={"update_path": "a", "reference_path": "b", "output_path": "c"},
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()["detail"]["code"], "PERMISSION_DENIED")
        self.client.headers.clear()
        self.client.headers.update(auth_headers_for_permissions(self.client, WEEKLY_PRICE_TEST_PERMISSIONS))


class WeeklyQuotePersistenceApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.original_db_dir = store.DB_DIR
        cls.original_db_path = store.DB_PATH
        store.close_connection()
        store.DB_DIR = cls.temp_dir.name
        store.DB_PATH = os.path.join(cls.temp_dir.name, "weekly-quote-persistence-test.db")
        store._connection = None
        store.init_database()
        cls.client_ctx = TestClient(app)
        cls.client = cls.client_ctx.__enter__()
        cls.client.headers.update(auth_headers_for_permissions(cls.client, WEEKLY_PRICE_TEST_PERMISSIONS))

    @classmethod
    def tearDownClass(cls):
        cls.client_ctx.__exit__(None, None, None)
        store.close_connection()
        store.DB_DIR = cls.original_db_dir
        store.DB_PATH = cls.original_db_path
        store._connection = None
        cls.temp_dir.cleanup()

    def test_save_and_list_batch(self):
        save_resp = self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄",
            "quote_date": "2026-05-07",
            "entries": [
                {"name": "大白菜", "unit": "斤", "price": 0.8},
                {"name": "西红柿", "unit": "斤", "price": 2.5},
            ],
        })
        self.assertEqual(save_resp.status_code, 200)
        self.assertTrue(save_resp.json()["success"])
        batch = save_resp.json()["batch"]
        self.assertEqual(batch["supplier"], "勾庄")
        self.assertEqual(len(batch["entries"]), 2)

        list_resp = self.client.get("/api/weekly-price/summary/batches?supplier=勾庄")
        self.assertEqual(list_resp.status_code, 200)
        self.assertEqual(len(list_resp.json()["batches"]), 1)

    def test_weekly_summary_highest_price(self):
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄", "quote_date": "2026-05-06",
            "entries": [
                {"name": "大白菜", "unit": "斤", "price": 0.9},
            ],
        })
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄", "quote_date": "2026-05-07",
            "entries": [
                {"name": "大白菜", "unit": "斤", "price": 0.8},
            ],
        })
        summary_resp = self.client.post("/api/weekly-price/summary/weekly", json={
            "supplier": "勾庄", "date": "2026-05-07",
        })
        self.assertEqual(summary_resp.status_code, 200)
        items = summary_resp.json()["summary_items"]
        cabbage = next((i for i in items if i["name"] == "大白菜"), None)
        self.assertIsNotNone(cabbage)
        self.assertEqual(cabbage["summary_price"], 0.9)

    def test_weekly_summary_uses_average_rule_for_lixiang(self):
        save_resp = self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "理想", "quote_date": "2026-07-06",
            "entries": [
                {"name": "青菜", "unit": "斤", "price": 1.0},
                {"name": "青菜", "unit": "斤", "price": 1.2},
            ],
        })
        self.assertEqual(save_resp.status_code, 200)

        summary_resp = self.client.post("/api/weekly-price/summary/weekly", json={
            "supplier": "理想", "date": "2026-07-06",
        })

        self.assertEqual(summary_resp.status_code, 200)
        payload = summary_resp.json()
        self.assertEqual(payload["batch_count"], 1)
        self.assertEqual(payload["entry_count"], 2)
        items = payload["summary_items"]
        greens = next((i for i in items if i["name"] == "青菜"), None)
        self.assertIsNotNone(greens)
        self.assertEqual(greens["summary_price"], 1.1)

    def test_save_rejects_supplier_weekly_batch_limit(self):
        first_resp = self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "理想", "quote_date": "2026-07-13",
            "entries": [{"name": "青菜", "unit": "斤", "price": 1.0}],
        })
        self.assertEqual(first_resp.status_code, 200)

        second_resp = self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "理想", "quote_date": "2026-07-14",
            "entries": [{"name": "白菜", "unit": "斤", "price": 1.2}],
        })

        self.assertEqual(second_resp.status_code, 400)
        self.assertIn("理想 本周最多只允许 1 个批次", second_resp.json()["detail"])

    def test_import_upload_parses_entries_without_saving(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "quotes.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "酱菜"
            worksheet.append(["菜名", "单位", "单价"])
            worksheet.append(["萝卜干", "包", 5.2])
            worksheet.append(["黄瓜条", "包", 4.8])
            workbook.save(workbook_path)

            with workbook_path.open("rb") as source_file:
                response = self.client.post(
                    "/api/weekly-price/summary/import/upload",
                    data={"supplier": "酱菜", "quote_date": "2026-07-20"},
                    files={
                        "source_file": (
                            "quotes.xlsx",
                            source_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )

        self.assertEqual(response.status_code, 200)
        batch = response.json()["batch"]
        self.assertEqual(batch["supplier"], "酱菜")
        self.assertEqual(batch["quote_date"], "2026-07-20")
        self.assertEqual(len(batch["entries"]), 2)

        list_resp = self.client.get("/api/weekly-price/summary/batches?supplier=酱菜")
        self.assertFalse(
            any(batch["quote_date"] == "2026-07-20" for batch in list_resp.json()["batches"])
        )

    def test_preview_validation_error_returns_400(self):
        response = self.client.post("/api/weekly-price/summary/preview", json={"batches": []})

        self.assertEqual(response.status_code, 400)
        self.assertIn("当前没有可汇总的报价批次", response.json()["detail"])

    def test_export_upload_returns_downloadable_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "summary-template.xlsx"
            workbook = Workbook()
            workbook.active.title = "勾庄模板"
            for sheet_name in ["理想模板", "刘慧模板", "酱菜模板", "豆制品模板"]:
                workbook.create_sheet(sheet_name)
            workbook.save(template_path)

            batches = [
                {
                    "supplier": "勾庄",
                    "quote_date": "2026-07-20",
                    "entries": [{"name": "白菜", "unit": "斤", "price": 2.5}],
                },
                {
                    "supplier": "理想",
                    "quote_date": "2026-07-20",
                    "entries": [{"name": "青菜", "unit": "斤", "price": 1.2}],
                },
            ]

            with template_path.open("rb") as workbook_file:
                response = self.client.post(
                    "/api/weekly-price/summary/export/upload",
                    data={"batches_json": json.dumps(batches, ensure_ascii=False)},
                    files={
                        "workbook_file": (
                            "summary-template.xlsx",
                            workbook_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment;", response.headers["content-disposition"])
        self.assertGreater(len(response.content), 0)

    def test_week_overview_returns_fixed_supplier_summaries_and_batches(self):
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄", "quote_date": "2026-08-03",
            "entries": [{"name": "白菜", "unit": "斤", "price": 2.0}],
        })
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄", "quote_date": "2026-08-04",
            "entries": [{"name": "白菜", "unit": "斤", "price": 2.6}],
        })
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "理想", "quote_date": "2026-08-03",
            "entries": [
                {"name": "青菜", "unit": "斤", "price": 1.0},
                {"name": "青菜", "unit": "斤", "price": 1.2},
            ],
        })

        response = self.client.get("/api/weekly-price/summary/week?date=2026-08-05")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["week_start"], "2026-08-03")
        self.assertEqual(payload["week_end"], "2026-08-09")
        self.assertEqual(
            [item["supplier"] for item in payload["suppliers"][:5]],
            ["勾庄", "理想", "刘慧", "酱菜", "豆制品"],
        )
        self.assertEqual(payload["total_batches"], 3)
        self.assertEqual(payload["total_entries"], 4)
        self.assertEqual(payload["total_summary_items"], 2)

        gouzhuang = payload["suppliers"][0]
        self.assertEqual(gouzhuang["limit"], 7)
        self.assertEqual(gouzhuang["batch_count"], 2)
        self.assertEqual(len(gouzhuang["batches"]), 2)
        cabbage = next(item for item in gouzhuang["summary_items"] if item["name"] == "白菜")
        self.assertEqual(cabbage["summary_price"], 2.6)

        lixiang = payload["suppliers"][1]
        self.assertEqual(lixiang["limit"], 1)
        self.assertEqual(lixiang["entry_count"], 2)
        greens = next(item for item in lixiang["summary_items"] if item["name"] == "青菜")
        self.assertEqual(greens["summary_price"], 1.1)

        liuhui = payload["suppliers"][2]
        self.assertEqual(liuhui["limit"], 1)
        self.assertEqual(liuhui["batches"], [])
        self.assertEqual(liuhui["summary_items"], [])

    def test_week_export_upload_uses_saved_week_data(self):
        save_resp = self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "豆制品", "quote_date": "2026-08-10",
            "entries": [{"name": "老豆腐", "unit": "斤", "price": 2.5}],
        })
        self.assertEqual(save_resp.status_code, 200)

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "summary-template.xlsx"
            workbook = Workbook()
            workbook.active.title = "勾庄模板"
            for sheet_name in ["理想模板", "刘慧模板", "酱菜模板", "豆制品模板"]:
                workbook.create_sheet(sheet_name)
            workbook.save(template_path)

            with template_path.open("rb") as workbook_file:
                response = self.client.post(
                    "/api/weekly-price/summary/export/week/upload",
                    data={"date": "2026-08-12"},
                    files={
                        "workbook_file": (
                            "summary-template.xlsx",
                            workbook_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment;", response.headers["content-disposition"])
        self.assertGreater(len(response.content), 0)

    def test_empty_week_export_upload_returns_400(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "summary-template.xlsx"
            workbook = Workbook()
            workbook.save(template_path)

            with template_path.open("rb") as workbook_file:
                response = self.client.post(
                    "/api/weekly-price/summary/export/week/upload",
                    data={"date": "2026-09-21"},
                    files={
                        "workbook_file": (
                            "summary-template.xlsx",
                            workbook_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )

        self.assertEqual(response.status_code, 400)
        self.assertIn("当前没有可汇总的报价批次", response.json()["detail"])

    def test_delete_batch(self):
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄", "quote_date": "2026-05-07",
            "entries": [
                {"name": "测试菜", "unit": "斤", "price": 1.0},
            ],
        })
        del_resp = self.client.post("/api/weekly-price/summary/delete", json={
            "supplier": "勾庄", "quote_date": "2026-05-07",
        })
        self.assertEqual(del_resp.status_code, 200)
        self.assertTrue(del_resp.json()["success"])

        list_resp = self.client.get("/api/weekly-price/summary/batches?supplier=勾庄")
        self.assertEqual(len(list_resp.json()["batches"]), 0)

    def test_list_suppliers(self):
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄", "quote_date": "2026-05-07",
            "entries": [{"name": "大白菜", "unit": "斤", "price": 0.8}],
        })
        self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "豆制品", "quote_date": "2026-05-07",
            "entries": [{"name": "老豆腐", "unit": "斤", "price": 2.5}],
        })
        resp = self.client.get("/api/weekly-price/summary/suppliers")
        self.assertEqual(resp.status_code, 200)
        suppliers = resp.json()["suppliers"]
        self.assertIn("勾庄", suppliers)
        self.assertIn("豆制品", suppliers)

    def test_summary_options_custom_supplier_units_rules_limits_and_export(self):
        options_resp = self.client.get("/api/weekly-price/summary/options")
        self.assertEqual(options_resp.status_code, 200)
        options = options_resp.json()
        self.assertEqual(
            [item["name"] for item in options["suppliers"][:5]],
            ["勾庄", "理想", "刘慧", "酱菜", "豆制品"],
        )
        self.assertIn("斤", [item["name"] for item in options["measure_units"]])

        supplier_resp = self.client.post("/api/weekly-price/summary/suppliers", json={
            "name": "自采API",
            "weekly_batch_limit": 2,
            "summary_rule": "average",
        })
        self.assertEqual(supplier_resp.status_code, 200)
        self.assertEqual(supplier_resp.json()["supplier"]["name"], "自采API")
        self.assertEqual(supplier_resp.json()["supplier"]["summary_rule"], "average")

        unit_resp = self.client.post("/api/weekly-price/summary/measure-units", json={"name": "盒_API测试"})
        self.assertEqual(unit_resp.status_code, 200)
        self.assertEqual(unit_resp.json()["measure_unit"]["name"], "盒_API测试")

        with tempfile.TemporaryDirectory() as tmpdir:
            import_path = Path(tmpdir) / "custom-import.xlsx"
            import_workbook = Workbook()
            import_sheet = import_workbook.active
            import_sheet.title = "自采API报价"
            import_sheet.append(["菜名", "单位", "单价"])
            import_sheet.append(["萝卜", "袋", 3.4])
            import_workbook.save(import_path)

            with import_path.open("rb") as source_file:
                import_resp = self.client.post(
                    "/api/weekly-price/summary/import/upload",
                    data={"supplier": "自采API", "quote_date": "2026-10-05"},
                    files={
                        "source_file": (
                            "custom-import.xlsx",
                            source_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )

        self.assertEqual(import_resp.status_code, 200)
        self.assertEqual(import_resp.json()["batch"]["entries"][0]["name"], "萝卜")

        for date, price in [("2026-10-05", 2.0), ("2026-10-06", 2.4)]:
            save_resp = self.client.post("/api/weekly-price/summary/save", json={
                "supplier": "自采API",
                "quote_date": date,
                "entries": [{"name": "青椒", "unit": "盒", "price": price}],
            })
            self.assertEqual(save_resp.status_code, 200)

        over_limit_resp = self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "自采API",
            "quote_date": "2026-10-07",
            "entries": [{"name": "白菜", "unit": "盒", "price": 1.2}],
        })
        self.assertEqual(over_limit_resp.status_code, 400)
        self.assertIn("自采API 本周最多只允许 2 个批次", over_limit_resp.json()["detail"])

        overview_resp = self.client.get("/api/weekly-price/summary/week?date=2026-10-06")
        self.assertEqual(overview_resp.status_code, 200)
        supplier = next(item for item in overview_resp.json()["suppliers"] if item["supplier"] == "自采API")
        self.assertEqual(supplier["limit"], 2)
        self.assertEqual(supplier["summary_rule"], "average")
        pepper = next(item for item in supplier["summary_items"] if item["name"] == "青椒")
        self.assertEqual(pepper["summary_price"], 2.2)

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "summary-template.xlsx"
            workbook = Workbook()
            workbook.active.title = "说明"
            workbook.save(template_path)

            with template_path.open("rb") as workbook_file:
                export_resp = self.client.post(
                    "/api/weekly-price/summary/export/week/upload",
                    data={"date": "2026-10-06"},
                    files={
                        "workbook_file": (
                            "summary-template.xlsx",
                            workbook_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )

        self.assertEqual(export_resp.status_code, 200)
        exported = load_workbook(BytesIO(export_resp.content))
        self.assertIn("自采API", exported.sheetnames)
        self.assertEqual(exported["自采API"]["A2"].value, "青椒")
        self.assertEqual(exported["自采API"]["B2"].value, "盒")
        self.assertEqual(exported["自采API"]["C2"].value, 2.2)

        refreshed_options = self.client.get("/api/weekly-price/summary/options").json()
        self.assertIn("盒", [item["name"] for item in refreshed_options["measure_units"]])

        save_new_unit_resp = self.client.post("/api/weekly-price/summary/save", json={
            "supplier": "勾庄",
            "quote_date": "2026-10-08",
            "entries": [{"name": "南瓜", "unit": "筐", "price": 12.0}],
        })
        self.assertEqual(save_new_unit_resp.status_code, 200)
        refreshed_after_save = self.client.get("/api/weekly-price/summary/options").json()
        self.assertIn("筐", [item["name"] for item in refreshed_after_save["measure_units"]])

    def test_summary_option_mutation_requires_create_permission(self):
        self.client.headers.clear()
        self.client.headers.update(auth_headers_for_permissions(self.client, ["weekly_quote:view"]))
        response = self.client.post("/api/weekly-price/summary/suppliers", json={
            "name": "无权限供应商",
            "weekly_batch_limit": 7,
            "summary_rule": "highest",
        })

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()["detail"]["code"], "PERMISSION_DENIED")
        self.client.headers.clear()
        self.client.headers.update(auth_headers_for_permissions(self.client, WEEKLY_PRICE_TEST_PERMISSIONS))


if __name__ == "__main__":
    unittest.main()
