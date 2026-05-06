from __future__ import annotations

import csv
import io
import re
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula


_DATE_PATTERNS = [
    re.compile(r"(?P<year>\d{4})[年./:：-](?P<month>\d{1,2})[月./:：-](?P<day>\d{1,2})日?"),
    re.compile(r"(?P<month>\d{1,2})[月./:：-](?P<day>\d{1,2})日?"),
]
_NAME_SPLIT_RE = re.compile(r"[,，、;；|\r\n]+")


class MonthlyListParser:
    def parse_text(self, raw_text: str, month: str) -> dict:
        rows: list[tuple[int, list[Any], str]] = []
        text = raw_text or ""
        reader = csv.reader(io.StringIO(text), delimiter="\t") if "\t" in text else csv.reader(io.StringIO(text))

        for line_number, cells in enumerate(reader, start=1):
            if not any(str(cell or "").strip() for cell in cells):
                continue
            raw_line = "\t".join("" if cell is None else str(cell) for cell in cells)
            rows.append((line_number, cells, raw_line))
        return self.parse_rows(rows, month)

    def parse_file(self, file_path: Path, month: str) -> dict:
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return self._parse_xlsx(file_path, month)
        if suffix in {".csv", ".txt"}:
            return self.parse_text(self._read_text_file(file_path), month)
        raise ValueError("月度清单只支持 CSV、TXT 或 XLSX 文件")

    def parse_rows(self, rows: list[tuple[int, list[Any], str]], month: str) -> dict:
        month_year, month_value = self._parse_month(month)
        days_in_month = monthrange(month_year, month_value)[1]
        by_date: dict[str, list[str]] = {}
        errors: list[dict] = []

        if not rows:
            errors.append({"line": 1, "message": "月度清单为空，请按“第一行日期、每列一天”填写", "raw": ""})
            return self._build_response(by_date, errors)

        header_line, header_cells, header_raw = rows[0]
        if self._looks_like_legacy_rows(rows, month_year, month_value):
            errors.append({
                "line": header_line,
                "message": "月度清单格式已调整为第一行写日期、每列一天；请把日期放在 A1、B1、C1，下方填写对应日期的菜名",
                "raw": header_raw,
            })
            return self._build_response(by_date, errors)

        max_columns = max(len(cells) for _, cells, _ in rows)
        for column_index in range(max_columns):
            column_label = self._column_label(column_index)
            header_value = header_cells[column_index] if column_index < len(header_cells) else ""
            data_cells = [
                cells[column_index]
                for _, cells, _ in rows[1:]
                if column_index < len(cells) and str(cells[column_index] or "").strip()
            ]
            formula_day = column_index + 1 if self._is_supported_formula_header(header_value) else None
            if formula_day and formula_day > days_in_month:
                if data_cells:
                    errors.append({
                        "line": header_line,
                        "message": f"{month_year:04d}-{month_value:02d} 没有第 {formula_day} 天，第 {column_label} 列已跳过",
                        "raw": str(header_value or ""),
                    })
                continue

            parsed_date = self._normalize_header_date(header_value, month_year, month_value, column_index)

            if not parsed_date:
                if str(header_value or "").strip() or data_cells:
                    errors.append({
                        "line": header_line,
                        "message": f"第 {column_label} 列首行未识别到日期，已跳过该列",
                        "raw": str(header_value or f"第 {column_label} 列"),
                    })
                continue

            names = self._parse_names(data_cells)
            if not names:
                errors.append({
                    "line": header_line,
                    "message": f"第 {column_label} 列 {parsed_date} 未识别到菜名",
                    "raw": str(header_value or ""),
                })
                continue

            existing = by_date.setdefault(parsed_date, [])
            seen = {name.lower() for name in existing}
            for name in names:
                key = name.lower()
                if key not in seen:
                    seen.add(key)
                    existing.append(name)

        detected_month = ""
        if by_date:
            detected_month = min(by_date.keys())[:7]

        return self._build_response(by_date, errors, detected_month)

    def _parse_xlsx(self, file_path: Path, month: str) -> dict:
        value_workbook = load_workbook(file_path, read_only=True, data_only=True)
        formula_workbook = load_workbook(file_path, read_only=True, data_only=False)
        try:
            value_sheet = value_workbook.active
            formula_sheet = formula_workbook.active
            max_row = max(value_sheet.max_row or 0, formula_sheet.max_row or 0)
            max_column = max(value_sheet.max_column or 0, formula_sheet.max_column or 0)
            value_rows = value_sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True)
            formula_rows = formula_sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True)
            rows: list[tuple[int, list[Any], str]] = []
            effective_month = month
            for row_index, (value_row, formula_row) in enumerate(zip(value_rows, formula_rows), start=1):
                cells = [
                    self._choose_xlsx_cell_value(value, formula)
                    for value, formula in zip(value_row, formula_row)
                ]

                array_formula_text: str | None = None
                for item in formula_row:
                    if isinstance(item, ArrayFormula):
                        array_formula_text = item.text
                        break

                if array_formula_text is not None:
                    detected_fm = self._detect_formula_month(array_formula_text)
                    if detected_fm and not (month and month.strip()):
                        effective_month = f"{date.today().year:04d}-{detected_fm:02d}"

                    for ci, cell in enumerate(cells):
                        if isinstance(cell, ArrayFormula):
                            cells[ci] = array_formula_text
                        elif self._is_blank(cell):
                            cells[ci] = array_formula_text

                if not any(not self._is_blank(cell) for cell in cells):
                    continue
                rows.append((row_index, cells, "\t".join("" if cell is None else str(cell) for cell in cells)))
            return self.parse_rows(rows, effective_month)
        finally:
            value_workbook.close()
            formula_workbook.close()

    @staticmethod
    def _build_response(by_date: dict[str, list[str]], errors: list[dict], detected_month: str = "") -> dict:
        entries = [
            {"date": entry_date, "names": names}
            for entry_date, names in sorted(by_date.items())
        ]
        return {
            "success": bool(entries) and not errors,
            "entries": entries,
            "errors": errors,
            "detected_month": detected_month,
            "total_dates": len(entries),
            "total_names": sum(len(entry["names"]) for entry in entries),
            "message": f"已解析 {len(entries)} 天、{sum(len(entry['names']) for entry in entries)} 个菜名",
        }

    @staticmethod
    def _read_text_file(file_path: Path) -> str:
        data = file_path.read_bytes()
        for encoding in ("utf-8-sig", "gbk", "utf-8"):
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    @staticmethod
    def _parse_month(month: str) -> tuple[int, int]:
        from datetime import date
        month = (month or "").strip()
        if not month:
            today = date.today()
            return today.year, today.month
        match = re.fullmatch(r"\s*(\d{4})-(\d{1,2})\s*", month)
        if not match:
            raise ValueError("月份必须使用 YYYY-MM 格式")
        year = int(match.group(1))
        month_value = int(match.group(2))
        if not 1 <= month_value <= 12:
            raise ValueError("月份必须在 1 到 12 之间")
        return year, month_value

    @staticmethod
    def _looks_like_header(cells: list[Any]) -> bool:
        text = " ".join("" if cell is None else str(cell) for cell in cells).lower()
        return ("日期" in text or "date" in text) and ("菜" in text or "品种" in text or "name" in text)

    def _looks_like_legacy_rows(self, rows: list[tuple[int, list[Any], str]], year: int, month: int) -> bool:
        if not rows:
            return False
        if self._looks_like_header(rows[0][1]):
            return True
        first_row_has_date = any(self._normalize_date(cell, year, month) for cell in rows[0][1])
        if first_row_has_date:
            return False
        for _, cells, _ in rows[1:]:
            if len(cells) >= 2 and any(self._normalize_date(cell, year, month) for cell in cells[:2]):
                return True
        return False

    @staticmethod
    def _column_label(index: int) -> str:
        label = ""
        current = index + 1
        while current:
            current, remainder = divmod(current - 1, 26)
            label = chr(65 + remainder) + label
        return label

    def _normalize_header_date(self, value: Any, default_year: int, default_month: int, column_index: int) -> str | None:
        parsed = self._normalize_date(value, default_year, default_month)
        if parsed:
            return parsed
        if not self._is_supported_formula_header(value):
            return None
        try:
            return date(default_year, default_month, column_index + 1).isoformat()
        except ValueError:
            return None

    def _normalize_date(self, value: Any, default_year: int, default_month: int) -> str | None:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

        text = str(value or "").strip()
        if not text:
            return None

        for pattern in _DATE_PATTERNS:
            match = pattern.search(text)
            if not match:
                continue
            year = int(match.groupdict().get("year") or default_year)
            month = int(match.groupdict().get("month") or default_month)
            day = int(match.group("day"))
            try:
                return date(year, month, day).isoformat()
            except ValueError:
                return None
        return None

    @staticmethod
    def _choose_xlsx_cell_value(value: Any, formula: Any) -> Any:
        if isinstance(formula, ArrayFormula):
            if not MonthlyListParser._is_blank(value):
                return value
            return formula.text
        if not MonthlyListParser._is_blank(value):
            return value
        if not MonthlyListParser._is_blank(formula):
            return formula
        return value

    @staticmethod
    def _is_blank(value: Any) -> bool:
        return value is None or str(value).strip() == ""

    @staticmethod
    def _detect_formula_month(formula_text: str) -> int | None:
        m = re.search(r'DATE\(YEAR\(TODAY\(\)\),\s*(\d+),\s*\d+\)', formula_text)
        if m:
            return int(m.group(1))
        return None

    @staticmethod
    def _is_supported_formula_header(value: Any) -> bool:
        text = str(value or "").strip().upper()
        return (
            text.startswith("=")
            and ("COLUMN(" in text or "SEQUENCE(" in text)
            and ("EOMONTH(" in text or "DATE(" in text)
            and ("TODAY(" in text or "YEAR(" in text or "MONTH(" in text)
        )

    @staticmethod
    def _parse_names(values: list[Any]) -> list[str]:
        names: list[str] = []
        seen: set[str] = set()
        for value in values:
            text = str(value or "").strip()
            if not text:
                continue
            text = re.sub(r"^(菜名|品种|名称)\s*[:：]\s*", "", text)
            parts = [part.strip() for part in _NAME_SPLIT_RE.split(text)]
            for part in parts:
                if not part:
                    continue
                key = part.lower()
                if key not in seen:
                    seen.add(key)
                    names.append(part)
        return names
