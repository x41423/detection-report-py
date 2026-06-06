import tempfile
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi import File, Form, UploadFile
from starlette.concurrency import run_in_threadpool

from backend.auth.dependencies import require_permission
from backend.api.upload_utils import (
    build_download_response,
    parse_json_form_value,
    save_upload,
)
from backend.models.schemas import (
    WeeklyPriceAliasDeleteRequest,
    WeeklyPriceAliasListResponse,
    WeeklyPriceAliasUpsertRequest,
    WeeklyPriceExecuteRequest,
    WeeklyPriceExecuteResponse,
    WeeklyPricePasteRequest,
    WeeklyPricePreviewRequest,
    WeeklyPricePreviewResponse,
    WeeklyQuoteDeleteRequest,
    WeeklyQuoteEntryInput,
    WeeklyQuoteExportRequest,
    WeeklyQuoteExportResponse,
    WeeklyQuoteImportRequest,
    WeeklyQuoteImportResponse,
    WeeklyQuoteMeasureUnitCreateRequest,
    WeeklyQuoteMeasureUnitCreateResponse,
    WeeklyQuotePreviewRequest,
    WeeklyQuotePreviewResponse,
    WeeklyQuoteSaveRequest,
    WeeklyQuoteSupplierCreateRequest,
    WeeklyQuoteSupplierCreateResponse,
    WeeklyQuoteSummaryOptionsResponse,
    WeeklyQuoteWeekOverviewResponse,
    WeeklyQuoteWeekSummaryRequest,
    WeeklyQuoteWeekSummaryResponse,
)
from backend.services.weekly_price_service import WeeklyPriceService
from backend.services.weekly_quote_summary_service import WeeklyQuoteSummaryService

router = APIRouter()
service = WeeklyPriceService()
summary_service = WeeklyQuoteSummaryService()


def _build_weekly_price_upload_output_name(update_file_name: str | None) -> str:
    stem = Path(str(update_file_name or "")).stem or "weekly-price"
    return f"{stem}_weekly_updated.xlsx"


def _raise_bad_request(exc: Exception) -> None:
    raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/preview", response_model=WeeklyPricePreviewResponse, dependencies=[Depends(require_permission("weekly_quote:view"))])
def preview_weekly_price(req: WeeklyPricePreviewRequest):
    try:
        result = service.preview(
            update_path=req.update_path,
            reference_path=req.reference_path,
        )
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return WeeklyPricePreviewResponse(**result)


@router.post(
    "/preview/upload",
    response_model=WeeklyPricePreviewResponse,
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
async def preview_weekly_price_upload(
    update_file: UploadFile = File(...),
    reference_file: UploadFile = File(...),
):
    with tempfile.TemporaryDirectory(prefix="weekly-price-preview-") as tmpdir:
        root = Path(tmpdir)
        update_path = await save_upload(update_file, root / "inputs", "update")
        reference_path = await save_upload(reference_file, root / "inputs", "reference")
        try:
            result = await run_in_threadpool(
                service.preview,
                str(update_path),
                str(reference_path),
            )
        except (FileNotFoundError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    return WeeklyPricePreviewResponse(**result)


@router.post(
    "/preview/templates",
    response_model=WeeklyPricePreviewResponse,
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
async def preview_from_templates():
    """Preview using stored templates from MinIO (no file upload needed)."""
    with tempfile.TemporaryDirectory(prefix="wp-preview-tpl-") as tmpdir:
        root = Path(tmpdir)
        (root / "inputs").mkdir(parents=True)
        update_path = root / "inputs" / TEMPLATE_OBJECTS["update"]
        ref_path = root / "inputs" / TEMPLATE_OBJECTS["reference"]
        try:
            if is_minio_enabled():
                for key, path in [("update", update_path), ("reference", ref_path)]:
                    resp = storage_service.client.get_object(MINIO_BUCKET, f"{TEMPLATE_PREFIX}/{TEMPLATE_OBJECTS[key]}")
                    path.write_bytes(resp.read())
                    resp.close()
                    resp.release_conn()
            else:
                for key, path in [("update", update_path), ("reference", ref_path)]:
                    src = Path(f"data/templates/{TEMPLATE_OBJECTS[key]}")
                    if not src.exists():
                        raise HTTPException(status_code=404, detail=f"{key} 模板尚未上传")
                    path.write_bytes(src.read_bytes())
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=404, detail="模板尚未上传，请先上传两个Excel文件")
        try:
            result = await run_in_threadpool(service.preview, str(update_path), str(ref_path))
        except (FileNotFoundError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    return WeeklyPricePreviewResponse(**result)


@router.post("/execute", response_model=WeeklyPriceExecuteResponse, dependencies=[Depends(require_permission("weekly_quote:update"))])
def execute_weekly_price(req: WeeklyPriceExecuteRequest):
    try:
        result = service.execute(
            update_path=req.update_path,
            reference_path=req.reference_path,
            output_path=req.output_path,
        )
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return WeeklyPriceExecuteResponse(**result)


@router.post("/execute/upload", dependencies=[Depends(require_permission("weekly_quote:update"))])
async def execute_weekly_price_upload(
    update_file: UploadFile = File(...),
    reference_file: UploadFile = File(...),
):
    with tempfile.TemporaryDirectory(prefix="weekly-price-execute-") as tmpdir:
        root = Path(tmpdir)
        update_path = await save_upload(update_file, root / "inputs", "update")
        reference_path = await save_upload(reference_file, root / "inputs", "reference")
        output_path = root / "output" / _build_weekly_price_upload_output_name(update_file.filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            result = await run_in_threadpool(
                service.execute,
                str(update_path),
                str(reference_path),
                str(output_path),
            )
        except (FileNotFoundError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        generated_path = Path(result["output_path"])
        if not generated_path.exists():
            raise HTTPException(status_code=500, detail="周报价更新已结束，但没有生成输出文件")

        return build_download_response(
            generated_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            extra_headers={
                "X-Operation-Message": result.get("message", "周报价更新已完成"),
                "X-Matched-Count": str(result.get("matched_count", 0)),
                "X-Updated-Count": str(result.get("updated_count", 0)),
            },
        )


@router.get("/aliases", response_model=WeeklyPriceAliasListResponse, dependencies=[Depends(require_permission("weekly_quote:aliases"))])
def list_weekly_price_aliases():
    result = service.list_aliases()
    return WeeklyPriceAliasListResponse(**result)


@router.put("/aliases", response_model=WeeklyPriceAliasListResponse, dependencies=[Depends(require_permission("weekly_quote:aliases"))])
def upsert_weekly_price_aliases(req: WeeklyPriceAliasUpsertRequest):
    result = service.upsert_aliases(req.mappings)
    return WeeklyPriceAliasListResponse(**result)


@router.delete("/aliases", response_model=WeeklyPriceAliasListResponse, dependencies=[Depends(require_permission("weekly_quote:aliases"))])
def delete_weekly_price_alias(req: WeeklyPriceAliasDeleteRequest):
    result = service.delete_alias(req.source_name)
    return WeeklyPriceAliasListResponse(**result)


@router.post(
    "/preview/paste",
    response_model=WeeklyPricePreviewResponse,
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
async def preview_weekly_price_paste(req: WeeklyPricePasteRequest):
    """粘贴模式预检：从粘贴的菜名和价格列表匹配模板。"""
    try:
        result = await run_in_threadpool(
            service.preview_from_paste,
            req.names,
            req.prices,
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    return WeeklyPricePreviewResponse(**result)


@router.post(
    "/execute/paste",
    dependencies=[Depends(require_permission("weekly_quote:update"))],
)
async def execute_weekly_price_paste(req: WeeklyPricePasteRequest):
    """粘贴模式执行：从粘贴的菜名和价格列表匹配模板并写入结果。"""
    with tempfile.TemporaryDirectory(prefix="weekly-price-paste-execute-") as tmpdir:
        output_path = Path(tmpdir) / "weekly_price_updated.xlsx"

        try:
            result = await run_in_threadpool(
                service.execute_from_paste,
                req.names,
                req.prices,
                str(output_path),
            )
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

        generated_path = Path(result["output_path"])
        if not generated_path.exists():
            raise HTTPException(status_code=500, detail="执行完成，但未生成输出文件")

        return build_download_response(
            generated_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            extra_headers={
                "X-Operation-Message": result.get("message", "执行完成"),
                "X-Matched-Count": str(result.get("matched_count", 0)),
                "X-Updated-Count": str(result.get("updated_count", 0)),
            },
        )


@router.post(
    "/import-reference",
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
async def import_reference_excel(file: UploadFile = File(...)):
    """从参考报价表 Excel 文件中读取菜名和价格。"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="请选择文件")
    
    ext = Path(file.filename).suffix.lower()
    if ext not in ('.xlsx', '.xls', '.xlsm'):
        raise HTTPException(status_code=400, detail="仅支持 Excel 文件格式")
    
    content_bytes = await file.read()
    
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.load_workbook(BytesIO(content_bytes), data_only=True)
        ws = wb.active
        
        names = []
        prices = []
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            
            # 处理菜名（将单元格内换行替换为空格）
            raw_name = str(row[0]).strip()
            name = raw_name.replace('\n', ' ').replace('\r', ' ')
            if not name:
                continue
            
            # 处理价格
            price = ''
            if len(row) > 1 and row[1] is not None:
                price = str(row[1]).strip()
            
            names.append(name)
            prices.append(price)
        
        wb.close()
        
        return {
            "success": True,
            "names": names,
            "prices": prices,
            "count": len(names),
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 Excel 失败：{str(exc)}")


@router.get(
    "/summary/options",
    response_model=WeeklyQuoteSummaryOptionsResponse,
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
def get_weekly_quote_summary_options():
    result = summary_service.get_options()
    return WeeklyQuoteSummaryOptionsResponse(**result)


@router.post(
    "/summary/suppliers",
    response_model=WeeklyQuoteSupplierCreateResponse,
    dependencies=[Depends(require_permission("weekly_quote:create"))],
)
def create_weekly_quote_supplier(req: WeeklyQuoteSupplierCreateRequest):
    try:
        result = summary_service.create_supplier(
            name=req.name,
            weekly_batch_limit=req.weekly_batch_limit,
            summary_rule=req.summary_rule,
        )
    except ValueError as exc:
        _raise_bad_request(exc)
    return WeeklyQuoteSupplierCreateResponse(**result)


@router.post(
    "/summary/measure-units",
    response_model=WeeklyQuoteMeasureUnitCreateResponse,
    dependencies=[Depends(require_permission("weekly_quote:create"))],
)
def create_weekly_quote_measure_unit(req: WeeklyQuoteMeasureUnitCreateRequest):
    try:
        result = summary_service.create_measure_unit(req.name)
    except ValueError as exc:
        _raise_bad_request(exc)
    return WeeklyQuoteMeasureUnitCreateResponse(**result)


@router.post(
    "/summary/import",
    response_model=WeeklyQuoteImportResponse,
    dependencies=[Depends(require_permission("weekly_quote:create"))],
)
def import_weekly_quote_summary(req: WeeklyQuoteImportRequest):
    try:
        result = summary_service.import_batch(
            supplier=req.supplier,
            quote_date=req.quote_date,
            source_path=req.source_path,
        )
    except (FileNotFoundError, ValueError) as exc:
        _raise_bad_request(exc)
    return WeeklyQuoteImportResponse(**result)


@router.post(
    "/summary/import/upload",
    response_model=WeeklyQuoteImportResponse,
    dependencies=[Depends(require_permission("weekly_quote:create"))],
)
async def import_weekly_quote_summary_upload(
    supplier: str = Form(...),
    quote_date: str = Form(...),
    source_file: UploadFile = File(...),
):
    with tempfile.TemporaryDirectory(prefix="weekly-quote-import-") as tmpdir:
        source_path = await save_upload(source_file, Path(tmpdir) / "inputs", "summary-import")
        try:
            result = await run_in_threadpool(
                summary_service.import_batch,
                supplier,
                quote_date,
                str(source_path),
            )
        except (FileNotFoundError, ValueError) as exc:
            _raise_bad_request(exc)
    return WeeklyQuoteImportResponse(**result)


@router.post(
    "/summary/preview",
    response_model=WeeklyQuotePreviewResponse,
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
def preview_weekly_quote_summary_data(req: WeeklyQuotePreviewRequest):
    try:
        result = summary_service.preview(
            batches=[batch.model_dump() for batch in req.batches],
        )
    except ValueError as exc:
        _raise_bad_request(exc)
    return WeeklyQuotePreviewResponse(**result)


@router.post(
    "/summary/export",
    response_model=WeeklyQuoteExportResponse,
    dependencies=[Depends(require_permission("weekly_quote:export"))],
)
def export_weekly_quote_summary_data(req: WeeklyQuoteExportRequest):
    try:
        result = summary_service.export(
            workbook_path=req.workbook_path,
            batches=[batch.model_dump() for batch in req.batches],
        )
    except (FileNotFoundError, PermissionError, ValueError) as exc:
        _raise_bad_request(exc)
    return WeeklyQuoteExportResponse(**result)


@router.post("/summary/export/upload", dependencies=[Depends(require_permission("weekly_quote:export"))])
async def export_weekly_quote_summary_data_upload(
    batches_json: str = Form(...),
    workbook_file: UploadFile = File(...),
):
    batches = parse_json_form_value(
        batches_json,
        field_name="batches_json",
        expected_type=list,
    )
    with tempfile.TemporaryDirectory(prefix="weekly-quote-export-") as tmpdir:
        root = Path(tmpdir)
        workbook_path = await save_upload(workbook_file, root / "inputs", "summary-template")
        try:
            result = await run_in_threadpool(
                summary_service.export,
                str(workbook_path),
                batches,
            )
        except (FileNotFoundError, PermissionError, ValueError) as exc:
            _raise_bad_request(exc)

        generated_path = Path(result["workbook_path"])
        if not generated_path.exists():
            raise HTTPException(status_code=500, detail="周报价汇总导出已结束，但没有生成输出文件")

        media_type = (
            "application/vnd.ms-excel.sheet.macroEnabled.12"
            if generated_path.suffix.lower() == ".xlsm"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return build_download_response(
            generated_path,
            media_type=media_type,
            extra_headers={"X-Operation-Message": result.get("message", "周报价汇总导出已完成")},
        )


@router.post("/summary/export/week/upload", dependencies=[Depends(require_permission("weekly_quote:export"))])
async def export_weekly_quote_summary_week_upload(
    date: str = Form(...),
    workbook_file: UploadFile = File(...),
):
    with tempfile.TemporaryDirectory(prefix="weekly-quote-week-export-") as tmpdir:
        root = Path(tmpdir)
        workbook_path = await save_upload(workbook_file, root / "inputs", "summary-template")
        try:
            result = await run_in_threadpool(
                summary_service.export_week,
                str(workbook_path),
                date,
            )
        except (FileNotFoundError, PermissionError, ValueError) as exc:
            _raise_bad_request(exc)

        generated_path = Path(result["workbook_path"])
        if not generated_path.exists():
            raise HTTPException(status_code=500, detail="周报价汇总导出已结束，但没有生成输出文件")

        media_type = (
            "application/vnd.ms-excel.sheet.macroEnabled.12"
            if generated_path.suffix.lower() == ".xlsm"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return build_download_response(
            generated_path,
            media_type=media_type,
            extra_headers={"X-Operation-Message": result.get("message", "周报价汇总导出已完成")},
        )


@router.post("/summary/save", dependencies=[Depends(require_permission("weekly_quote:create"))])
def save_quote_batch(req: WeeklyQuoteSaveRequest):
    svc = WeeklyQuoteSummaryService()
    try:
        return svc.save_manual_batch(
            req.supplier, req.quote_date,
            [e.model_dump() for e in req.entries],
            req.source_label,
        )
    except ValueError as exc:
        _raise_bad_request(exc)


@router.get("/summary/batches", dependencies=[Depends(require_permission("weekly_quote:view"))])
def list_quote_batches(supplier: str):
    svc = WeeklyQuoteSummaryService()
    return svc.list_saved_batches(supplier)


@router.post("/summary/delete", dependencies=[Depends(require_permission("weekly_quote:delete"))])
def delete_quote_batch(req: WeeklyQuoteDeleteRequest):
    svc = WeeklyQuoteSummaryService()
    return svc.delete_batch(req.supplier, req.quote_date)


@router.get(
    "/summary/week",
    response_model=WeeklyQuoteWeekOverviewResponse,
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
def weekly_quote_week_overview(date: str):
    svc = WeeklyQuoteSummaryService()
    try:
        return svc.get_week_overview(date)
    except ValueError as exc:
        _raise_bad_request(exc)


@router.post(
    "/summary/weekly",
    response_model=WeeklyQuoteWeekSummaryResponse,
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
def weekly_quote_summary(req: WeeklyQuoteWeekSummaryRequest):
    svc = WeeklyQuoteSummaryService()
    try:
        return svc.get_weekly_summary(req.supplier, req.date)
    except ValueError as exc:
        _raise_bad_request(exc)


@router.get("/summary/suppliers", dependencies=[Depends(require_permission("weekly_quote:view"))])
def list_quote_suppliers():
    svc = WeeklyQuoteSummaryService()
    return svc.get_all_suppliers()


# ==================================================================
# Template storage (MinIO) — edit uploaded Excel directly
# ==================================================================

from io import BytesIO
from backend.services.storage_service import is_minio_enabled, storage_service, MINIO_BUCKET

TEMPLATE_PREFIX = "weekly-templates"
TEMPLATE_OBJECTS = {"update": "update-template.xlsx", "reference": "reference-template.xlsx"}


def _read_excel_grid(file_bytes: bytes) -> list[list[str]]:
    """Parse Excel bytes into a 2D array of cell values."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    grid: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        grid.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return grid


def _write_excel_grid(original_bytes: bytes, grid: list[list[str]]) -> bytes:
    """Overwrite cell values in an existing Excel template, preserving formats."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(original_bytes))
    ws = wb.active
    for r_idx, row in enumerate(grid, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.read()


@router.post(
    "/template/{tmpl_type}",
    dependencies=[Depends(require_permission("weekly_quote:update"))],
)
async def upload_template(tmpl_type: str, file: UploadFile = File(...)):
    """Upload template to MinIO. tmpl_type = 'update' | 'reference'."""
    if tmpl_type not in ("update", "reference"):
        raise HTTPException(status_code=400, detail="type must be 'update' or 'reference'")
    content = await file.read()
    if is_minio_enabled():
        storage_service.upload_file(content, f"{TEMPLATE_PREFIX}/{TEMPLATE_OBJECTS[tmpl_type]}",
                                     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # 始终保存本地备份（MinIO 不可用时使用）
    (Path("data/templates")).mkdir(parents=True, exist_ok=True)
    Path(f"data/templates/{TEMPLATE_OBJECTS[tmpl_type]}").write_bytes(content)
    return {"success": True, "message": "模板已上传"}


@router.post(
    "/template/{tmpl_type}/from-path",
    dependencies=[Depends(require_permission("weekly_quote:update"))],
)
async def upload_template_from_path(tmpl_type: str, file_path: str = Form(...)):
    """Upload template from server path. tmpl_type = 'update' | 'reference'."""
    if tmpl_type not in ("update", "reference"):
        raise HTTPException(status_code=400, detail="type must be 'update' or 'reference'")
    source = Path(file_path.strip())
    if not source.is_file():
        raise HTTPException(status_code=400, detail=f"模板文件不存在: {file_path}")
    content = source.read_bytes()
    if is_minio_enabled():
        storage_service.upload_file(content, f"{TEMPLATE_PREFIX}/{TEMPLATE_OBJECTS[tmpl_type]}",
                                     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # 始终保存本地备份
    (Path("data/templates")).mkdir(parents=True, exist_ok=True)
    Path(f"data/templates/{TEMPLATE_OBJECTS[tmpl_type]}").write_bytes(content)
    return {"success": True, "message": "模板已上传"}


@router.get(
    "/templates",
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
def list_templates():
    """List stored template files with metadata."""
    result = {}
    for key in ("update", "reference"):
        obj = TEMPLATE_OBJECTS[key]
        if is_minio_enabled():
            try:
                stat = storage_service.client.stat_object(MINIO_BUCKET, f"{TEMPLATE_PREFIX}/{obj}")
                result[key] = {"name": obj, "size": stat.size, "updated": str(stat.last_modified)}
            except Exception as e:
                result[key] = {"name": obj, "error": str(e)}
        else:
            p = Path(f"data/templates/{obj}")
            if p.exists():
                result[key] = {"name": obj, "size": p.stat().st_size,
                               "updated": str(p.stat().st_mtime)}
            else:
                result[key] = None
    return {"success": True, "templates": result}


@router.get(
    "/template/{tmpl_type}/read",
    dependencies=[Depends(require_permission("weekly_quote:view"))],
)
def read_template(tmpl_type: str):
    """Read template Excel content as a 2D JSON grid."""
    if tmpl_type not in ("update", "reference"):
        raise HTTPException(status_code=400, detail="type must be 'update' or 'reference'")
    obj = TEMPLATE_OBJECTS[tmpl_type]
    if is_minio_enabled():
        try:
            resp = storage_service.client.get_object(MINIO_BUCKET, f"{TEMPLATE_PREFIX}/{obj}")
            content = resp.read()
            resp.close()
            resp.release_conn()
        except Exception as e:
            raise HTTPException(status_code=404, detail=f"读取模板失败: {e}")
    else:
        p = Path(f"data/templates/{obj}")
        if not p.exists():
            raise HTTPException(status_code=404, detail="模板尚未上传")
        content = p.read_bytes()
    grid = _read_excel_grid(content)
    return {"success": True, "grid": grid, "rows": len(grid), "cols": len(grid[0]) if grid else 0}


@router.put(
    "/template/{tmpl_type}",
    dependencies=[Depends(require_permission("weekly_quote:update"))],
)
def save_template(tmpl_type: str, grid: list[list[str]]):
    """Save edited grid back to template Excel on MinIO."""
    if tmpl_type not in ("update", "reference"):
        raise HTTPException(status_code=400, detail="type must be 'update' or 'reference'")
    obj = TEMPLATE_OBJECTS[tmpl_type]
    # Read original to preserve formatting
    if is_minio_enabled():
        try:
            resp = storage_service.client.get_object(MINIO_BUCKET, f"{TEMPLATE_PREFIX}/{obj}")
            original = resp.read()
            resp.close()
            resp.release_conn()
        except Exception:
            raise HTTPException(status_code=404, detail="模板尚未上传")
    else:
        p = Path(f"data/templates/{obj}")
        if not p.exists():
            raise HTTPException(status_code=404, detail="模板尚未上传")
        original = p.read_bytes()
    new_bytes = _write_excel_grid(original, grid)
    if is_minio_enabled():
        storage_service.upload_file(new_bytes, f"{TEMPLATE_PREFIX}/{obj}",
                                     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        Path(f"data/templates/{obj}").write_bytes(new_bytes)
    return {"success": True, "message": "模板已保存"}
